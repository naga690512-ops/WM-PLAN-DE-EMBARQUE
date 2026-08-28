"""
label_engine.py
Motor de procesamiento: toma un export crudo del portal (CSV/XLSX), arma la
estructura de cajas, y genera el workbook Excel + Etiquetas 1/2/3 en ZPL y PDF.

Toda la lógica de aquí fue validada el 18-ago-2026 contra impresión real:
- El bit del QR (GRF hex) usa la polaridad correcta (1=negro, 0=blanco).
- El espaciado de Etiqueta 2 replica el archivo de referencia (1 a 5 productos).
- El layout del PDF no tiene textos encimados (validado visualmente).
"""
import re
import json
from io import BytesIO
from collections import defaultdict

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Utilidades de estilo Excel (versión inline, sin depender de un módulo externo)
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def apply_header_style(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def freeze_header_row(ws, row=2):
    ws.freeze_panes = f"A{row}"


def apply_alternating_fill(ws, start_row, end_row):
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for cell in ws[r]:
                cell.fill = ALT_FILL


def format_column_as_text(ws, col_letter):
    for cell in ws[col_letter]:
        cell.number_format = "@"


def autofit_columns(ws, max_width=45):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column_letter
            widths[col] = max(widths.get(col, 0), len(str(cell.value)))
    for col, w in widths.items():
        ws.column_dimensions[col].width = min(w + 2, max_width)


def format_ean_value(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ---------------------------------------------------------------------------
# Mapeo flexible de columnas
# ---------------------------------------------------------------------------

CAMPOS_REQUERIDOS = {
    "oc": ["Orden Compra", "Núm. Pedido", "Numero Pedido", "OC", "Pedido"],
    "cadena": ["Cadena", "CADENA"],
    "ean": ["Ean/Upc", "EAN-13", "EAN", "UPC", "Ean"],
    "estilo": ["Estilo", "Modelo"],
    "talla": ["Talla", "TALLA"],
    "color": ["Color", "COLOR"],
    "descripcion": ["Descripcion", "Descripción", "DESCRIPCION"],
    "departamento": ["depto", "Departamento", "DEPTO", "Depto"],
    "tienda": ["Tienda", "TIENDA"],
    "cantidad": ["Cantidad", "CANTIDAD", "Cant"],
    "cedis": ["Cedis", "CEDIS"],
    "proveedor": ["# Proveedor", "Proveedor", "No. Proveedor"],
}


def sugerir_mapeo(columnas_archivo):
    """Devuelve {campo: columna_sugerida_o_None} comparando contra alias conocidos."""
    mapeo = {}
    cols_lower = {c.lower().strip(): c for c in columnas_archivo if c}
    for campo, alias in CAMPOS_REQUERIDOS.items():
        encontrado = None
        for a in alias:
            if a.lower() in cols_lower:
                encontrado = cols_lower[a.lower()]
                break
        mapeo[campo] = encontrado
    return mapeo


def extraer_cedis_de_cadena(cadena):
    m = re.match(r"\s*(\d+)", str(cadena))
    return m.group(1) if m else ""


def leer_filas_crudas(file_bytes, filename, mapeo, cedis_desde_cadena,
                       departamento_fijo, descripcion_fija):
    """Lee el archivo crudo (csv o xlsx) y devuelve una lista de dicts normalizados."""
    if filename.lower().endswith(".csv"):
        import csv
        import io
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        raw_rows = list(reader)
    else:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [c.value for c in ws[1]]
        raw_rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            raw_rows.append(dict(zip(headers, r)))

    rows = []
    for d in raw_rows:
        oc_val = d.get(mapeo["oc"])
        if oc_val is None or str(oc_val).strip() == "":
            continue
        if cedis_desde_cadena:
            cedis = extraer_cedis_de_cadena(d.get(mapeo["cadena"], ""))
        else:
            cedis = format_ean_value(d.get(mapeo.get("cedis")))

        descripcion = (
            str(d.get(mapeo["descripcion"])).strip()
            if mapeo.get("descripcion") and d.get(mapeo["descripcion"]) not in (None, "")
            else descripcion_fija
        )
        departamento = (
            str(d.get(mapeo["departamento"])).strip()
            if mapeo.get("departamento") and d.get(mapeo["departamento"]) not in (None, "")
            else departamento_fijo
        )

        rows.append({
            "Núm. Pedido": int(float(oc_val)),
            "Cadena": d.get(mapeo.get("cadena"), ""),
            "Cedis": cedis,
            "Proveedor": d.get(mapeo["proveedor"]) if mapeo.get("proveedor") else None,
            "EAN-13": format_ean_value(d.get(mapeo["ean"])),
            "Estilo": d.get(mapeo["estilo"]),
            "Talla": str(d.get(mapeo["talla"])),
            "Color": d.get(mapeo["color"]),
            "Descripcion": descripcion,
            "Departamento": departamento,
            "Tienda": int(float(d.get(mapeo["tienda"]))),
            "Cantidad": int(float(d.get(mapeo["cantidad"]))),
        })
    return rows


# ---------------------------------------------------------------------------
# Construcción de cajas (modalidad: 1 por tienda, o máximo de piezas por caja)
# ---------------------------------------------------------------------------

def construir_cajas(rows, modalidad, max_piezas=None):
    """modalidad: '1_por_tienda' o 'max_piezas'. Devuelve lista de boxes (dicts)."""
    grouped = defaultdict(list)
    for r in rows:
        grouped[(r["Núm. Pedido"], r["Tienda"])].append(r)

    boxes = []
    for (pedido, tienda), items in sorted(grouped.items()):
        cedis = items[0]["Cedis"]
        departamento = items[0]["Departamento"]
        descripcion = items[0]["Descripcion"]

        if modalidad == "1_por_tienda" or not max_piezas:
            chunks = [items]
        else:
            chunks, current, current_total = [], [], 0
            for it in items:
                if current and current_total + it["Cantidad"] > max_piezas:
                    chunks.append(current)
                    current, current_total = [], 0
                current.append(it)
                current_total += it["Cantidad"]
            if current:
                chunks.append(current)

        cadena = items[0]["Cadena"]
        n_cajas = len(chunks)
        for idx, chunk in enumerate(chunks, start=1):
            estilos = sorted(set(x["Estilo"] for x in chunk))
            total_piezas = sum(x["Cantidad"] for x in chunk)
            productos = [{"ean": x["EAN-13"], "talla": x["Talla"], "color": x["Color"],
                          "cant": str(x["Cantidad"]), "estilo": x["Estilo"]} for x in chunk]
            boxes.append({
                "oc": pedido, "cadena": cadena, "cedis": cedis, "tienda": tienda,
                "departamento": departamento, "descripcion": descripcion,
                "estilos": estilos, "total_piezas": total_piezas,
                "productos": productos, "caja_txt": f"{idx} de {n_cajas}",
            })
    return boxes


# ---------------------------------------------------------------------------
# Workbook Excel — replica el formato exacto usado con Label Matrix
# (colores, fuente y estructura de columnas validados contra envios_WM.xlsx,
# el archivo de referencia de producción, el 19-ago-2026)
# ---------------------------------------------------------------------------

REF_HEADER_FILL = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
REF_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
REF_DATA_FONT = Font(name="Arial", size=10, bold=False)
REF_ALIGN = Alignment(horizontal="center", vertical="center")


def _ean_a_numero(v):
    """Convierte a int si es un EAN puramente numérico (así lo espera Label Matrix)."""
    s = str(v).strip()
    return int(s) if s.isdigit() else s


def _escribir_hoja(wb, nombre, headers, filas, ean_cols=()):
    ws = wb.create_sheet(nombre)
    ws.append(headers)
    for r in filas:
        ws.append(r)
    for cell in ws[1]:
        cell.fill = REF_HEADER_FILL
        cell.font = REF_HEADER_FONT
        cell.alignment = REF_ALIGN
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = REF_DATA_FONT
            cell.alignment = REF_ALIGN
    ws.freeze_panes = "A2"
    for col_name in ean_cols:
        if col_name in headers:
            letter = get_column_letter(headers.index(col_name) + 1)
            for cell in ws[letter][1:]:
                cell.number_format = "0"
    autofit_columns(ws)
    return ws


def construir_workbook(rows, boxes):
    wb = Workbook()
    wb.remove(wb.active)

    tiene_proveedor = any(r.get("Proveedor") not in (None, "") for r in rows)
    tiene_desc = any(r.get("Descripcion") not in (None, "") for r in rows)
    tiene_depto = any(r.get("Departamento") not in (None, "") for r in rows)

    # ---- Detalle Envíos ----
    detalle_headers = ["Núm. Pedido", "Cadena"]
    if tiene_proveedor:
        detalle_headers.append("# Proveedor")
    detalle_headers += ["EAN-13", "Estilo", "Talla", "Color"]
    if tiene_desc:
        detalle_headers.append("Descripcion")
    if tiene_depto:
        detalle_headers.append("Departamento")
    detalle_headers += ["Tienda", "Cantidad"]

    filas_detalle = []
    for r in rows:
        fila = [r["Núm. Pedido"], r["Cadena"]]
        if tiene_proveedor:
            fila.append(r.get("Proveedor"))
        fila += [_ean_a_numero(r["EAN-13"]), r["Estilo"], r["Talla"], r["Color"]]
        if tiene_desc:
            fila.append(r.get("Descripcion"))
        if tiene_depto:
            fila.append(r.get("Departamento"))
        fila += [r["Tienda"], r["Cantidad"]]
        filas_detalle.append(fila)
    _escribir_hoja(wb, "Detalle Envíos", detalle_headers, filas_detalle, ean_cols=["EAN-13"])

    # ---- Resumen por Tienda ----
    res_headers = ["Núm. Pedido", "Cadena", "Tienda", "Total Piezas", "Núm. Estilos", "Núm. EANs"]
    grouped = defaultdict(list)
    for r in rows:
        grouped[(r["Núm. Pedido"], r["Tienda"])].append(r)
    filas_res = []
    for (pedido, tienda), items in sorted(grouped.items()):
        total_piezas = sum(x["Cantidad"] for x in items)
        n_estilos = len(set(x["Estilo"] for x in items))
        n_eans = len(set(x["EAN-13"] for x in items))
        filas_res.append([pedido, items[0]["Cadena"], tienda, total_piezas, n_estilos, n_eans])
    _escribir_hoja(wb, "Resumen por Tienda", res_headers, filas_res)

    # ---- Lineal x Tienda (agrupado por caja real x estilo, como Label Matrix) ----
    grupos_lineal = []
    for b in boxes:
        sub_estilo = defaultdict(list)
        for p in b["productos"]:
            sub_estilo[p["estilo"]].append(p)
        for estilo, items in sub_estilo.items():
            grupos_lineal.append((b, estilo, items))
    max_p_lineal = max((len(items) for _, _, items in grupos_lineal), default=0)

    lin_headers = ["Núm. Pedido", "Cadena", "Cedis", "Tienda", "Estilo", "Total Pzas x Estilo"]
    for i in range(1, max_p_lineal + 1):
        lin_headers += [f"EAN-13 P{i}", f"Talla P{i}", f"Color P{i}", f"Cant P{i}"]
    lin_headers += ["EAN / Cantidad", "Número de Cajas x tienda"]

    filas_lin = []
    for b, estilo, items in grupos_lineal:
        total_pzas = sum(int(p["cant"]) for p in items)
        fila = [b["oc"], b["cadena"], b["cedis"], b["tienda"], estilo, total_pzas]
        qr_lines = []
        for p in items:
            fila += [_ean_a_numero(p["ean"]), p["talla"], p["color"], int(p["cant"])]
            qr_lines += [p["ean"], p["cant"]]
        for _ in range(max_p_lineal - len(items)):
            fila += [None, None, None, None]
        fila.append("\n".join(qr_lines))
        fila.append(b["caja_txt"])
        filas_lin.append(fila)
    ean_cols_lineal = [f"EAN-13 P{i}" for i in range(1, max_p_lineal + 1)]
    _escribir_hoja(wb, "Lineal x Tienda", lin_headers, filas_lin, ean_cols=ean_cols_lineal)

    # ---- Una hoja por Orden de Compra (formato ancho, 1 fila = 1 caja) ----
    pedidos = sorted(set(b["oc"] for b in boxes))
    for pedido in pedidos:
        sub = [b for b in boxes if b["oc"] == pedido]
        max_p = max((len(b["productos"]) for b in sub), default=0)

        po_headers = ["Núm. Pedido", "Cadena", "Cedis", "Tienda", "Total Piezas Tienda",
                      "Piezas x Caja", "Caja"]
        for i in range(1, max_p + 1):
            po_headers += [f"EAN-13 P{i}", f"Talla P{i}", f"Color P{i}", f"Cant P{i}"]
        po_headers += ["EAN / Cantidad", "Núm. EANs x Caja"]

        total_tienda = defaultdict(int)
        for b in sub:
            total_tienda[b["tienda"]] += b["total_piezas"]

        filas_po = []
        for b in sorted(sub, key=lambda x: x["tienda"]):
            fila = [b["oc"], b["cadena"], b["cedis"], b["tienda"], total_tienda[b["tienda"]],
                    b["total_piezas"], b["caja_txt"]]
            qr_lines = []
            for p in b["productos"]:
                fila += [_ean_a_numero(p["ean"]), p["talla"], p["color"], int(p["cant"])]
                qr_lines += [p["ean"], p["cant"]]
            for _ in range(max_p - len(b["productos"])):
                fila += [None, None, None, None]
            fila.append("\n".join(qr_lines))
            fila.append(len(b["productos"]))
            filas_po.append(fila)

        ean_cols_po = [f"EAN-13 P{i}" for i in range(1, max_p + 1)]
        _escribir_hoja(wb, str(pedido), po_headers, filas_po, ean_cols=ean_cols_po)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# ZPL: Etiquetas 1, 2 y 3 (con el fix de polaridad de bits del QR)
# ---------------------------------------------------------------------------

DPMM = 203 / 25.4


def d(mm_val):
    return int(mm_val * DPMM)


def qr_to_grf_hex(content, size_mm=44):
    import qrcode
    from PIL import Image

    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=4, border=2)
    qr.add_data(content)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("1")

    target = d(size_mm)
    img_resized = img.resize((target, target), Image.NEAREST)
    w, h = img_resized.size
    pixels = img_resized.load()

    bytes_per_row = (w + 7) // 8
    padding = bytes_per_row * 8 - w
    grf_hex_parts = []
    for y in range(h):
        row = 0
        for x in range(w):
            bit = 1 if pixels[x, y] == 0 else 0  # 1=negro (imprime), 0=blanco -- CORRECTO, validado 18-ago-2026
            row = (row << 1) | bit
        row <<= padding
        for i in range(bytes_per_row - 1, -1, -1):
            grf_hex_parts.append(f"{(row >> (i * 8)) & 0xFF:02X}")
    grf_hex = "".join(grf_hex_parts)
    total_bytes = bytes_per_row * h
    return total_bytes, bytes_per_row, grf_hex


def wrap_modelos(modelos, max_chars=42, max_lines=2):
    lines, current = [], ""
    for m in modelos:
        piece = m if not current else ", " + m
        if len(current) + len(piece) > max_chars and current:
            lines.append(current)
            current = m
        else:
            current += piece
    if current:
        lines.append(current)
    if len(lines) > max_lines:
        head = lines[:max_lines - 1]
        tail = ", ".join(lines[max_lines - 1:])
        lines = head + [tail]
    return lines


def etiqueta1_zpl(temporada, cedis, oc, modelos_lines, tienda, piezas_caja, caja_txt,
                   caja_barcode, departamento, descripcion):
    lines = ["^XA", "^PW812", "^LL1218", "^CI28", "^LH0,0",
             "^FO23,15^GB351,255,255^FS",
             f"^FO39,31^FR^CF0,120^FD{temporada}^FS",
             "^FO439,23^CF0,55^FDRFID^FS",
             "^FO23,303^CF0,46^FDCEDIS:^FS",
             f"^FO215,303^CF0,46^FD{cedis}^FS",
             f"^FO455,303^BY2,2,111^BCN,111,N,N,N^FD{cedis}^FS",
             f"^FO455,421^CF0,28^FD{cedis}^FS",
             "^FO23,471^CF0,46^FDOC:^FS",
             f"^FO431,471^BY2,2,111^BCN,111,N,N,N^FD{oc}^FS",
             f"^FO431,589^CF0,28^FD{oc}^FS",
             "^FO23,631^CF0,46^FDMODELOS:^FS"]
    y = 631
    for ml in modelos_lines:
        lines.append(f"^FO247,{y}^CF0,28^FD{ml}^FS")
        y += 47
    lines += ["^FO23,757^CF0,46^FDTIENDA DESTINO:^FS",
              f"^FO431,757^BY2,2,111^BCN,111,N,N,N^FD{tienda}^FS",
              f"^FO431,875^CF0,28^FD{tienda}^FS",
              f"^FO23,805^CF0,52^FD{tienda}^FS",
              "^FO23,879^CF0,46^FDDEPARTAMENTO:^FS",
              f"^FO367,879^CF0,46^FD{departamento}^FS",
              "^FO23,943^CF0,46^FDDESCRIPCION:^FS",
              f"^FO335,943^CF0,46^FD{descripcion}^FS",
              "^FO0,1007^GB812,2,2^FS",
              f"^FO15,1022^BY2,2,79^BCN,79,N,N,N^FD{piezas_caja}^FS",
              f"^FO15,1108^CF0,26^FD{piezas_caja}^FS",
              "^FO15,1140^CF0,20^FDTOTAL DE PIEZAS EN CAJA^FS",
              f"^FO421,1022^BY2,2,79^BCN,79,N,N,N^FD{caja_barcode}^FS",
              f"^FO421,1108^CF0,26^FD{caja_txt}^FS",
              "^FO421,1140^CF0,20^FDCONSECUTIVO DE CAJA^FS", "^XZ"]
    return "\n".join(lines)


LAYOUT_POR_N = {
    1: (504, None, 393), 2: (303, 401, 292), 3: (203, 301, 242),
    4: (143, 241, 212), 5: (109, 207, 195),
}


def etiqueta2_zpl(tienda, caja_txt, productos):
    pages = []
    for i in range(0, len(productos), 5):
        chunk = productos[i:i + 5]
        n = len(chunk)
        start, interval, sep_offset = LAYOUT_POR_N[n]
        lines = ["^XA", "^PW812", "^LL1218", "^CI28", "^LH0,0",
                 "^FO38,7^CF0,28^FDUPC^FS", "^FO477,7^CF0,28^FDDESC.^FS",
                 "^FO580,7^CF0,28^FDTALLA^FS", "^FO653,7^CF0,28^FDPIEZAS^FS",
                 "^FO0,63^GB812,2,2^FS", "^FO0,1107^GB812,2,2^FS",
                 f"^FO23,1130^CF0,32^FDTIENDA  {tienda}^FS",
                 f"^FO445,1130^CF0,32^FDCAJA  {caja_txt}^FS"]
        y = start
        for p in chunk:
            lines.append(f"^FO23,{y}^BY2,2,126^BCN,126,N,N,N^FD{p['ean']}^FS")
            lines.append(f"^FO23,{y + 137}^CF0,28^FD{p['ean']}^FS")
            lines.append(f"^FO477,{y + 77}^CF0,26^FD{p['color']}^FS")
            lines.append(f"^FO580,{y + 77}^CF0,28^FD{p['talla']}^FS")
            lines.append(f"^FO646,{y}^BY2,2,126^BCN,126,N,N,N^FD{p['cant']}^FS")
            lines.append(f"^FO646,{y + 137}^CF0,28^FD{p['cant']}^FS")
            lines.append(f"^FO0,{y + sep_offset}^GB812,1,1,B^FS")
            if interval:
                y += interval
        lines.append("^XZ")
        pages.append("\n".join(lines))
    return "\n".join(pages)


def etiqueta3_zpl(empresa, cedis, oc, tienda, num_upcs, piezas, caja_txt, qr_content, grf_name):
    total_bytes, bytes_per_row, grf_hex = qr_to_grf_hex(qr_content, size_mm=44)
    W, H = d(101.6), d(152.4)
    qr_x, qr_y = d(46), d(3)
    caja_disp = caja_txt.upper()

    lines = ["^XA", f"^PW{W}", f"^LL{H}", "^CI28", "^LH0,0",
             f"~DGR:{grf_name},{total_bytes},{bytes_per_row},{grf_hex}",
             f"^FO{d(3)},{d(3)}^CF0,28^FD{empresa}^FS",
             f"^FO{qr_x},{qr_y}^XGR:{grf_name},1,1^FS",
             f"^FO{d(3)},{d(16)}^CF0,40^FDCEDIS^FS",
             f"^FO{d(25)},{d(16)}^CF0,40^FD{cedis}^FS",
             f"^FO{d(3)},{d(28)}^CF0,40^FDTIENDA^FS",
             f"^FO{d(3)},{d(36)}^BY2,2,{d(14)}^BCN,{d(14)},N,N,N^FD{tienda}^FS",
             f"^FO{d(3)},{d(54)}^CF0,55^FD{tienda}^FS",
             f"^FO{d(46)},{d(50)}^CF0,36^FDUPC  {num_upcs}^FS",
             f"^FO{d(46)},{d(59)}^CF0,36^FDPIEZAS EN LA CAJA  {piezas}^FS",
             f"^FO{d(46)},{d(68)}^CF0,26^FDCONSECUTIVO DE CAJA^FS",
             f"^FO{d(46)},{d(75)}^CF0,26^FDPOR TIENDA^FS",
             f"^FO{d(50)},{d(82)}^CF0,50^FD{caja_disp}^FS",
             f"^FO{d(3)},{d(68)}^CF0,36^FDORDEN DE COMPRA^FS",
             f"^FO{d(3)},{d(76)}^BY2,2,{d(14)}^BCN,{d(14)},N,N,N^FD{oc}^FS",
             f"^FO{d(3)},{d(94)}^CF0,43^FD{oc}^FS", "^XZ"]
    return "\n".join(lines)


def generar_zpl_para_oc(boxes, oc, temporada, empresa):
    out_blocks = []
    for b in [x for x in boxes if x["oc"] == oc]:
        cedis = b["cedis"]
        tienda = str(b["tienda"])
        piezas_caja = str(b["total_piezas"])
        caja_txt = b["caja_txt"]
        caja_barcode = caja_txt.replace(" ", "")
        productos = b["productos"]
        modelos_lines = wrap_modelos(b["estilos"])

        out_blocks.append(etiqueta1_zpl(temporada, cedis, oc, modelos_lines, tienda, piezas_caja,
                                         caja_txt, caja_barcode, b["departamento"], b["descripcion"]))
        out_blocks.append(etiqueta2_zpl(tienda, caja_txt, productos))

        qr_lines = []
        for p in productos:
            qr_lines += [p["ean"], p["cant"]]
        qr_content = "\n".join(qr_lines)
        grf_name = f"QR{tienda}_{caja_barcode}.GRF"
        out_blocks.append(etiqueta3_zpl(empresa, cedis, oc, tienda, len(productos), piezas_caja,
                                         caja_txt, qr_content, grf_name))
    return "\n".join(out_blocks) + "\n"


def generar_zpl_solo_qr_para_oc(boxes, oc, empresa):
    """Genera un ZPL que trae ÚNICAMENTE la Etiqueta 3 (QR) de cada caja del pedido."""
    out_blocks = []
    for b in [x for x in boxes if x["oc"] == oc]:
        cedis = b["cedis"]
        tienda = str(b["tienda"])
        piezas_caja = str(b["total_piezas"])
        caja_txt = b["caja_txt"]
        caja_barcode = caja_txt.replace(" ", "")
        productos = b["productos"]

        qr_lines = []
        for p in productos:
            qr_lines += [p["ean"], p["cant"]]
        qr_content = "\n".join(qr_lines)
        grf_name = f"QR{tienda}_{caja_barcode}.GRF"
        out_blocks.append(etiqueta3_zpl(empresa, cedis, oc, tienda, len(productos), piezas_caja,
                                         caja_txt, qr_content, grf_name))
    return "\n".join(out_blocks) + "\n"


# --- Etiqueta 3 + fotos por modelo (hasta 8, en cuadrícula automática) ---

FOTO_AREA_X_MM = 6
FOTO_AREA_Y_MM = 100
FOTO_AREA_W_MM = 89.6
FOTO_AREA_H_MM = 48
FOTO_PAD_MM = 1.5
MAX_FOTOS_POR_ETIQUETA = 8


def calcular_grid_fotos(n):
    """Devuelve (filas, columnas) para acomodar n fotos (1 a 8) en el área disponible."""
    n = max(1, min(n, MAX_FOTOS_POR_ETIQUETA))
    if n <= 4:
        return 1, n
    return 2, -(-n // 2)  # ceil(n/2) columnas


def foto_a_grf_hex(img, max_w_mm=FOTO_AREA_W_MM, max_h_mm=FOTO_AREA_H_MM):
    """Convierte una foto (PIL Image) a GRF hex con dithering (para verse bien
    en el térmico, a diferencia del QR que usa umbral duro). Devuelve también
    el tamaño final en px para poder centrarla en su área."""
    import numpy as np
    from PIL import Image

    img = img.convert("L")
    target_w_px = d(max_w_mm)
    target_h_px = d(max_h_mm)
    ratio = min(target_w_px / img.width, target_h_px / img.height)
    new_w, new_h = max(1, int(img.width * ratio)), max(1, int(img.height * ratio))
    img_resized = img.resize((new_w, new_h), Image.LANCZOS)
    bw = img_resized.convert("1")  # dithering Floyd-Steinberg por defecto

    arr = np.array(bw)
    black_mask = (arr == 0).astype(np.uint8)  # 0=negro en modo '1' de PIL
    packed = np.packbits(black_mask, axis=1, bitorder="big")
    bytes_per_row = packed.shape[1]
    hexdata = packed.tobytes().hex().upper()
    total_bytes = bytes_per_row * new_h
    return new_w, new_h, bytes_per_row, total_bytes, hexdata


def etiqueta3_multifoto_zpl(empresa, cedis, oc, tienda, num_upcs, piezas, caja_txt,
                             qr_content, grf_name, fotos, grf_prefix, cache_grf):
    """fotos: lista de (estilo, PIL.Image) ya en el orden a mostrar (máx 8).
    cache_grf: dict compartido entre cajas para no recodificar la misma foto
    al mismo tamaño de celda más de una vez."""
    base = etiqueta3_zpl(empresa, cedis, oc, tienda, num_upcs, piezas, caja_txt, qr_content, grf_name)
    base = base.rsplit("^XZ", 1)[0]

    if not fotos:
        return base + "^XZ"

    n = len(fotos)
    filas, cols = calcular_grid_fotos(n)
    cell_w_mm = FOTO_AREA_W_MM / cols
    cell_h_mm = FOTO_AREA_H_MM / filas
    area_w_px, area_h_px = d(cell_w_mm), d(cell_h_mm)

    bloques_foto = []
    for i, (estilo, img) in enumerate(fotos):
        fila_i, col_i = divmod(i, cols)
        cache_key = (estilo, round(cell_w_mm, 3), round(cell_h_mm, 3))
        if cache_key not in cache_grf:
            cache_grf[cache_key] = foto_a_grf_hex(img, cell_w_mm - FOTO_PAD_MM, cell_h_mm - FOTO_PAD_MM)
        fw, fh, fbpr, ftotal, fhex = cache_grf[cache_key]

        off_x = (area_w_px - fw) // 2
        off_y = (area_h_px - fh) // 2
        x = d(FOTO_AREA_X_MM + col_i * cell_w_mm) + off_x
        y = d(FOTO_AREA_Y_MM + fila_i * cell_h_mm) + off_y
        nombre = f"{grf_prefix}{i}.GRF"
        bloques_foto.append(f"~DGR:{nombre},{ftotal},{fbpr},{fhex}\n^FO{x},{y}^XGR:{nombre},1,1^FS\n")

    return base + "".join(bloques_foto) + "^XZ"


def generar_zpl_qr_fotos_para_oc(boxes, oc, empresa, fotos_por_estilo):
    """fotos_por_estilo: dict {estilo: PIL.Image}. Cada caja muestra las fotos
    de los modelos que trae (hasta 8), en el orden de su lista de estilos."""
    cache_grf = {}
    out_blocks = []
    for b in [x for x in boxes if x["oc"] == oc]:
        cedis = b["cedis"]
        tienda = str(b["tienda"])
        piezas_caja = str(b["total_piezas"])
        caja_txt = b["caja_txt"]
        caja_barcode = caja_txt.replace(" ", "")
        productos = b["productos"]

        qr_lines = []
        for p in productos:
            qr_lines += [p["ean"], p["cant"]]
        qr_content = "\n".join(qr_lines)
        grf_name = f"QR{tienda}_{caja_barcode}.GRF"

        fotos_caja = [(e, fotos_por_estilo[e]) for e in b["estilos"] if e in fotos_por_estilo]
        fotos_caja = fotos_caja[:MAX_FOTOS_POR_ETIQUETA]
        grf_prefix = f"FT{tienda}_{caja_barcode}_"
        out_blocks.append(etiqueta3_multifoto_zpl(empresa, cedis, oc, tienda, len(productos), piezas_caja,
                                                    caja_txt, qr_content, grf_name, fotos_caja, grf_prefix, cache_grf))
    return "\n".join(out_blocks) + "\n"


# ---------------------------------------------------------------------------
# PDF: Etiquetas 1, 2 y 3
# ---------------------------------------------------------------------------

from reportlab.lib.units import mm as MM
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.graphics.barcode import code128
from reportlab.lib.utils import ImageReader

PAGE_W, PAGE_H = 101.6 * MM, 152.4 * MM


def _y_top(dist_from_top_mm):
    return PAGE_H - dist_from_top_mm * MM


def _draw_barcode(c, data, x_mm, top_mm, height_mm, bar_width=0.2):
    bc = code128.Code128(str(data), barHeight=height_mm * MM, barWidth=bar_width * MM)
    bc.drawOn(c, x_mm * MM, _y_top(top_mm) - height_mm * MM)


def _qr_image_reader(content):
    import qrcode
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=8, border=2)
    qr.add_data(content)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return ImageReader(buf)


def _etiqueta1_pdf(c, temporada, cedis, oc, modelos, tienda, piezas, caja_txt, departamento, descripcion):
    c.setFillColorRGB(0, 0, 0)
    c.rect(3 * MM, _y_top(35), 44 * MM, 32 * MM, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 34)
    c.drawString(7 * MM, _y_top(28), temporada)
    c.setFillColorRGB(0, 0, 0)

    c.setFont("Helvetica-Bold", 16)
    c.drawString(56 * MM, _y_top(8), "RFID")

    c.setFont("Helvetica-Bold", 15)
    c.drawString(3 * MM, _y_top(42), "CEDIS:")
    cedis_x = 3 + c.stringWidth("CEDIS: ", "Helvetica-Bold", 15) / MM + 2
    c.drawString(cedis_x * MM, _y_top(42), str(cedis))
    _draw_barcode(c, cedis, 57, 38, 14)
    c.setFont("Helvetica", 8)
    c.drawString(57 * MM, _y_top(56), str(cedis))

    c.setFont("Helvetica-Bold", 15)
    c.drawString(3 * MM, _y_top(61), "OC:")
    _draw_barcode(c, oc, 54, 58, 14)
    c.setFont("Helvetica", 8)
    c.drawString(54 * MM, _y_top(76), str(oc))

    c.setFont("Helvetica-Bold", 15)
    c.drawString(3 * MM, _y_top(80), "MODELOS:")
    c.setFont("Helvetica-Bold", 9)
    y = 80
    modelos_x = 3 + c.stringWidth("MODELOS: ", "Helvetica-Bold", 15) / MM + 2
    for line in modelos:
        c.drawString(modelos_x * MM, _y_top(y), line)
        y += 6

    c.setFont("Helvetica-Bold", 15)
    c.drawString(3 * MM, _y_top(97), "TIENDA DESTINO:")
    _draw_barcode(c, tienda, 54, 93, 13)
    c.setFont("Helvetica", 8)
    c.drawString(54 * MM, _y_top(110), str(tienda))
    c.setFont("Helvetica-Bold", 17)
    c.drawString(3 * MM, _y_top(112), str(tienda))

    c.setFont("Helvetica-Bold", 15)
    c.drawString(3 * MM, _y_top(122), "DEPARTAMENTO:")
    depto_x = 3 + c.stringWidth("DEPARTAMENTO: ", "Helvetica-Bold", 15) / MM + 2
    c.drawString(depto_x * MM, _y_top(122), str(departamento))

    c.setFont("Helvetica-Bold", 15)
    c.drawString(3 * MM, _y_top(131), "DESCRIPCION:")
    desc_x = 3 + c.stringWidth("DESCRIPCION: ", "Helvetica-Bold", 15) / MM + 2
    c.drawString(desc_x * MM, _y_top(131), str(descripcion))

    c.line(0, _y_top(138), PAGE_W, _y_top(138))

    _draw_barcode(c, piezas, 2, 139, 7)
    c.setFont("Helvetica", 8)
    c.drawString(2 * MM, _y_top(148.5), str(piezas))
    c.setFont("Helvetica-Bold", 6)
    c.drawString(2 * MM, _y_top(151.5), "TOTAL DE PIEZAS EN CAJA")

    caja_barcode = caja_txt.replace(" ", "")
    _draw_barcode(c, caja_barcode, 53, 139, 7)
    c.setFont("Helvetica", 8)
    c.drawString(53 * MM, _y_top(148.5), caja_txt)
    c.setFont("Helvetica-Bold", 6)
    c.drawString(53 * MM, _y_top(151.5), "CONSECUTIVO DE CAJA")

    c.showPage()


def _etiqueta2_pdf(c, tienda, caja_txt, productos):
    for i in range(0, len(productos), 5):
        chunk = productos[i:i + 5]
        c.setFont("Helvetica-Bold", 9)
        c.drawString(3 * MM, _y_top(6), "UPC")
        c.drawString(41 * MM, _y_top(6), "DESC.")
        c.drawString(63 * MM, _y_top(6), "TALLA")
        c.drawString(80 * MM, _y_top(6), "PIEZAS")
        c.line(0, _y_top(8), PAGE_W, _y_top(8))

        row_h = 28
        y = 12
        for p in chunk:
            _draw_barcode(c, p["ean"], 3, y, 12, bar_width=0.15)
            c.setFont("Helvetica", 7)
            c.drawString(3 * MM, _y_top(y + 15), p["ean"])
            c.setFont("Helvetica-Bold", 9)
            c.drawString(41 * MM, _y_top(y + 9), p["color"])
            c.drawString(63 * MM, _y_top(y + 9), str(p["talla"]))
            _draw_barcode(c, p["cant"], 80, y, 12, bar_width=0.15)
            c.setFont("Helvetica", 7)
            c.drawString(80 * MM, _y_top(y + 15), str(p["cant"]))
            c.line(0, _y_top(y + row_h - 3), PAGE_W, _y_top(y + row_h - 3))
            y += row_h

        c.line(0, _y_top(142), PAGE_W, _y_top(142))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(3 * MM, _y_top(148), f"TIENDA  {tienda}")
        c.drawString(56 * MM, _y_top(148), f"CAJA  {caja_txt}")
        c.showPage()


def _etiqueta3_pdf_contenido(c, empresa, cedis, oc, tienda, num_upcs, piezas, caja_txt, qr_content):
    c.setFont("Helvetica-Bold", 8)
    c.drawString(3 * MM, _y_top(6), empresa)

    qr_img = _qr_image_reader(qr_content)
    qr_size = 44 * MM
    c.drawImage(qr_img, 46 * MM, _y_top(47), width=qr_size, height=qr_size)

    c.setFont("Helvetica-Bold", 13)
    c.drawString(3 * MM, _y_top(18), "CEDIS")
    c.drawString(20 * MM, _y_top(18), str(cedis))

    c.drawString(3 * MM, _y_top(30), "TIENDA")
    _draw_barcode(c, tienda, 3, 34, 14)
    c.setFont("Helvetica-Bold", 17)
    c.drawString(3 * MM, _y_top(56), str(tienda))

    c.setFont("Helvetica-Bold", 12)
    c.drawString(46 * MM, _y_top(52), f"UPC  {num_upcs}")
    c.drawString(46 * MM, _y_top(60), f"PIEZAS EN LA CAJA  {piezas}")
    c.setFont("Helvetica-Bold", 8)
    c.drawString(46 * MM, _y_top(68), "CONSECUTIVO DE CAJA")
    c.drawString(46 * MM, _y_top(73), "POR TIENDA")
    c.setFont("Helvetica-Bold", 15)
    c.drawString(50 * MM, _y_top(82), caja_txt.upper())

    c.setFont("Helvetica-Bold", 10)
    c.drawString(3 * MM, _y_top(70), "ORDEN DE COMPRA")
    _draw_barcode(c, oc, 3, 74, 14)
    c.setFont("Helvetica-Bold", 14)
    c.drawString(3 * MM, _y_top(96), str(oc))


def _etiqueta3_pdf(c, empresa, cedis, oc, tienda, num_upcs, piezas, caja_txt, qr_content):
    _etiqueta3_pdf_contenido(c, empresa, cedis, oc, tienda, num_upcs, piezas, caja_txt, qr_content)
    c.showPage()


def generar_pdf_para_oc(boxes, oc, temporada, empresa):
    buf = BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=(PAGE_W, PAGE_H))
    for b in [x for x in boxes if x["oc"] == oc]:
        cedis = b["cedis"]
        tienda = str(b["tienda"])
        piezas = str(b["total_piezas"])
        caja_txt = b["caja_txt"]
        productos = b["productos"]
        modelos_lines = wrap_modelos(b["estilos"])

        _etiqueta1_pdf(c, temporada, cedis, oc, modelos_lines, tienda, piezas, caja_txt,
                       b["departamento"], b["descripcion"])
        _etiqueta2_pdf(c, tienda, caja_txt, productos)

        qr_lines = []
        for p in productos:
            qr_lines += [p["ean"], p["cant"]]
        _etiqueta3_pdf(c, empresa, cedis, oc, tienda, len(productos), piezas, caja_txt, "\n".join(qr_lines))

    c.save()
    buf.seek(0)
    return buf


def generar_pdf_solo_qr_para_oc(boxes, oc, empresa):
    """Genera un PDF que trae ÚNICAMENTE la Etiqueta 3 (QR) de cada caja del pedido."""
    buf = BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=(PAGE_W, PAGE_H))
    for b in [x for x in boxes if x["oc"] == oc]:
        cedis = b["cedis"]
        tienda = str(b["tienda"])
        piezas = str(b["total_piezas"])
        caja_txt = b["caja_txt"]
        productos = b["productos"]

        qr_lines = []
        for p in productos:
            qr_lines += [p["ean"], p["cant"]]
        _etiqueta3_pdf(c, empresa, cedis, oc, tienda, len(productos), piezas, caja_txt, "\n".join(qr_lines))

    c.save()
    buf.seek(0)
    return buf


def _etiqueta3_multifoto_pdf(c, empresa, cedis, oc, tienda, num_upcs, piezas, caja_txt, qr_content, fotos_readers):
    """fotos_readers: lista de (estilo, ImageReader) ya en el orden a mostrar (máx 8)."""
    _etiqueta3_pdf_contenido(c, empresa, cedis, oc, tienda, num_upcs, piezas, caja_txt, qr_content)

    n = len(fotos_readers)
    if n > 0:
        filas, cols = calcular_grid_fotos(n)
        cell_w_mm = FOTO_AREA_W_MM / cols
        cell_h_mm = FOTO_AREA_H_MM / filas
        cell_w, cell_h = cell_w_mm * MM, cell_h_mm * MM

        for i, (estilo, foto_reader) in enumerate(fotos_readers):
            fila_i, col_i = divmod(i, cols)
            iw, ih = foto_reader.getSize()
            ratio = min((cell_w - FOTO_PAD_MM * MM) / iw, (cell_h - FOTO_PAD_MM * MM) / ih)
            new_w, new_h = iw * ratio, ih * ratio
            cell_x_mm = FOTO_AREA_X_MM + col_i * cell_w_mm
            cell_y_top_mm = FOTO_AREA_Y_MM + fila_i * cell_h_mm
            x = cell_x_mm * MM + (cell_w - new_w) / 2
            offset_y_mm = (cell_h_mm - new_h / MM) / 2
            y = _y_top(cell_y_top_mm + offset_y_mm) - new_h
            c.drawImage(foto_reader, x, y, width=new_w, height=new_h, preserveAspectRatio=True, mask="auto")

    c.showPage()


def generar_pdf_qr_fotos_para_oc(boxes, oc, empresa, fotos_por_estilo):
    """fotos_por_estilo: dict {estilo: bytes de la imagen}. Cada caja muestra
    las fotos de los modelos que trae (hasta 8)."""
    readers_por_estilo = {e: ImageReader(BytesIO(fb)) for e, fb in fotos_por_estilo.items()}

    buf = BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=(PAGE_W, PAGE_H))
    for b in [x for x in boxes if x["oc"] == oc]:
        cedis = b["cedis"]
        tienda = str(b["tienda"])
        piezas = str(b["total_piezas"])
        caja_txt = b["caja_txt"]
        productos = b["productos"]

        qr_lines = []
        for p in productos:
            qr_lines += [p["ean"], p["cant"]]

        fotos_caja = [(e, readers_por_estilo[e]) for e in b["estilos"] if e in readers_por_estilo]
        fotos_caja = fotos_caja[:MAX_FOTOS_POR_ETIQUETA]
        _etiqueta3_multifoto_pdf(c, empresa, cedis, oc, tienda, len(productos), piezas, caja_txt,
                                  "\n".join(qr_lines), fotos_caja)

    c.save()
    buf.seek(0)
    return buf
