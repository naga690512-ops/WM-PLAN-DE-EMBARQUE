"""
app.py — Plan de Embarque WM: automatiza Excel + Etiquetas 1/2/3 (ZPL y PDF)
a partir del export crudo del portal (CSV/XLSX).
"""
import streamlit as st
import label_engine as eng

st.set_page_config(page_title="Plan de Embarque WM", page_icon="📦", layout="wide")
st.title("📦 Plan de Embarque WM")
st.caption("Sube el export crudo del portal → Excel procesado + Etiquetas 1, 2 y 3 (ZPL y PDF).")

if "rows" not in st.session_state:
    st.session_state.rows = None
if "boxes" not in st.session_state:
    st.session_state.boxes = None

uploaded = st.file_uploader("Archivo del portal (CSV o XLSX)", type=["csv", "xlsx"])

if uploaded:
    file_bytes = uploaded.getvalue()

    # Detectar columnas disponibles
    if uploaded.name.lower().endswith(".csv"):
        import csv, io
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        columnas = next(reader)
    else:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
        columnas = [c.value for c in next(wb[wb.sheetnames[0]].iter_rows(min_row=1, max_row=1))]

    st.success(f"Archivo leído: {len(columnas)} columnas detectadas.")
    mapeo_sugerido = eng.sugerir_mapeo(columnas)

    st.subheader("1. Confirma las columnas")
    st.caption("Detecté esto automáticamente — ajusta si algo no cuadra.")

    col1, col2 = st.columns(2)
    opciones = ["(ninguna)"] + list(columnas)

    def selector(campo, label, columna_obj):
        idx = opciones.index(mapeo_sugerido[campo]) if mapeo_sugerido[campo] in opciones else 0
        return columna_obj.selectbox(label, opciones, index=idx, key=f"map_{campo}")

    with col1:
        col_oc = selector("oc", "Orden de Compra / Pedido", col1)
        col_ean = selector("ean", "EAN / UPC", col1)
        col_estilo = selector("estilo", "Estilo", col1)
        col_talla = selector("talla", "Talla", col1)
        col_color = selector("color", "Color", col1)
        col_proveedor = selector("proveedor", "# Proveedor (opcional)", col1)
    with col2:
        col_tienda = selector("tienda", "Tienda", col2)
        col_cantidad = selector("cantidad", "Cantidad", col2)
        col_cadena = selector("cadena", "Cadena (para extraer CEDIS)", col2)
        col_desc = selector("descripcion", "Descripción (opcional)", col2)
        col_depto = selector("departamento", "Departamento (opcional)", col2)

    st.subheader("2. Datos que no vienen en el archivo")
    c1, c2, c3 = st.columns(3)
    with c1:
        descripcion_fija = st.text_input(
            "Descripción fija", value="",
            disabled=(col_desc != "(ninguna)"),
            help="Se usa solo si no seleccionaste una columna de Descripción arriba.",
        )
    with c2:
        departamento_fijo = st.text_input(
            "Departamento fijo", value="",
            disabled=(col_depto != "(ninguna)"),
            help="Se usa solo si no seleccionaste una columna de Departamento arriba.",
        )
    with c3:
        temporada = st.text_input("Leyenda de temporada (Etiqueta 1)", value="T3")

    empresa = st.text_input("Empresa (Etiqueta 3)", value="NANCY GREEYS SA DE CV")

    st.subheader("3. Modalidad de cajas")
    modalidad_label = st.radio(
        "¿Cómo se arman las cajas por tienda?",
        ["1 caja por tienda (sin límite)", "Máximo de piezas por caja"],
        horizontal=True,
    )
    max_piezas = None
    if modalidad_label == "Máximo de piezas por caja":
        max_piezas = st.number_input("Máximo de piezas por caja", min_value=1, value=60, step=1)
    modalidad = "1_por_tienda" if modalidad_label.startswith("1 caja") else "max_piezas"

    faltan = [c for c, v in [("Orden de Compra", col_oc), ("EAN", col_ean), ("Estilo", col_estilo),
                              ("Talla", col_talla), ("Color", col_color), ("Tienda", col_tienda),
                              ("Cantidad", col_cantidad)] if v == "(ninguna)"]

    if faltan:
        st.warning(f"Faltan columnas obligatorias: {', '.join(faltan)}")
    else:
        if st.button("🚀 Procesar", type="primary"):
            mapeo = {
                "oc": col_oc, "ean": col_ean, "estilo": col_estilo, "talla": col_talla,
                "color": col_color, "tienda": col_tienda, "cantidad": col_cantidad,
                "cadena": None if col_cadena == "(ninguna)" else col_cadena,
                "descripcion": None if col_desc == "(ninguna)" else col_desc,
                "departamento": None if col_depto == "(ninguna)" else col_depto,
                "proveedor": None if col_proveedor == "(ninguna)" else col_proveedor,
            }
            cedis_desde_cadena = mapeo["cadena"] is not None

            with st.spinner("Leyendo archivo..."):
                rows = eng.leer_filas_crudas(file_bytes, uploaded.name, mapeo, cedis_desde_cadena,
                                              departamento_fijo, descripcion_fija)
            with st.spinner("Armando cajas..."):
                boxes = eng.construir_cajas(rows, modalidad, max_piezas)

            st.session_state.rows = rows
            st.session_state.boxes = boxes
            st.session_state.temporada = temporada
            st.session_state.empresa = empresa
            st.session_state.zpl_cache = {}
            st.session_state.pdf_cache = {}
            st.session_state.zpl_qr_cache = {}
            st.session_state.pdf_qr_cache = {}
            st.session_state.zpl_fotos_cache = {}
            st.session_state.pdf_fotos_cache = {}
            st.session_state.foto_asignaciones = {}
            st.session_state.fotos_acumuladas = {}
            st.session_state.excel_cache = None
            st.success(f"Listo: {len(rows)} líneas → {len(boxes)} cajas.")

if st.session_state.boxes:
    rows = st.session_state.rows
    boxes = st.session_state.boxes
    temporada = st.session_state.temporada
    empresa = st.session_state.empresa

    pedidos = sorted(set(b["oc"] for b in boxes))
    total_piezas = sum(b["total_piezas"] for b in boxes)

    st.divider()
    st.subheader("4. Resultados")
    m1, m2, m3 = st.columns(3)
    m1.metric("Pedidos", len(pedidos))
    m2.metric("Cajas", len(boxes))
    m3.metric("Piezas totales", total_piezas)

    with st.spinner("Generando Excel..."):
        if st.session_state.get("excel_cache") is None:
            st.session_state.excel_cache = eng.construir_workbook(rows, boxes)
    st.download_button("⬇️ Descargar Excel procesado", st.session_state.excel_cache,
                        file_name="Plan_de_Embarque.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.subheader("5. Fotos por modelo para Etiqueta QR + fotos (opcional)")
    st.caption("Nombra cada archivo con el código de Estilo exacto (ej. G026J04627.jpg). "
               "Cada caja mostrará las fotos de los modelos que trae, hasta 8, acomodadas en cuadrícula. "
               "Puedes subir en varias rondas — se van acumulando.")
    estilos_conocidos = sorted(set(e for b in boxes for e in b["estilos"]))
    fotos_subidas = st.file_uploader("Fotografías (una por modelo)", type=["jpg", "jpeg", "png"],
                                      accept_multiple_files=True, key="fotos_uploader")

    if "fotos_acumuladas" not in st.session_state:
        st.session_state.fotos_acumuladas = {}  # nombre -> bytes

    if fotos_subidas:
        for f in fotos_subidas:
            if f.name not in st.session_state.fotos_acumuladas:
                st.session_state.fotos_acumuladas[f.name] = f.getvalue()
                st.session_state.zpl_fotos_cache = {}
                st.session_state.pdf_fotos_cache = {}

    fotos_por_estilo = {}
    if st.session_state.fotos_acumuladas:
        nombres = list(st.session_state.fotos_acumuladas.keys())
        if len(nombres) > 8:
            st.warning(f"Tienes {len(nombres)} fotos acumuladas; solo se usarán las primeras 8. "
                       "Quita alguna abajo si necesitas otra combinación.")
        nombres = nombres[:8]

        if st.button("🗑️ Quitar todas las fotos"):
            st.session_state.fotos_acumuladas = {}
            st.session_state.foto_asignaciones = {}
            st.session_state.zpl_fotos_cache = {}
            st.session_state.pdf_fotos_cache = {}
            st.rerun()

        if "foto_asignaciones" not in st.session_state:
            st.session_state.foto_asignaciones = {}

        st.caption("Confirma o corrige a qué modelo corresponde cada foto:")
        cols_fotos = st.columns(min(len(nombres), 4))
        opciones_estilo = ["(ninguno)"] + estilos_conocidos
        for i, nombre in enumerate(nombres):
            with cols_fotos[i % len(cols_fotos)]:
                st.image(st.session_state.fotos_acumuladas[nombre], width=120)
                nombre_sin_ext = nombre.rsplit(".", 1)[0].strip()
                sugerido = nombre_sin_ext if nombre_sin_ext in estilos_conocidos else "(ninguno)"
                idx = opciones_estilo.index(st.session_state.foto_asignaciones.get(nombre, sugerido)) \
                    if st.session_state.foto_asignaciones.get(nombre, sugerido) in opciones_estilo else 0
                elegido = st.selectbox(nombre, opciones_estilo, index=idx, key=f"asign_{nombre}")
                st.session_state.foto_asignaciones[nombre] = elegido
                if elegido != "(ninguno)":
                    fotos_por_estilo[elegido] = st.session_state.fotos_acumuladas[nombre]

    st.session_state.fotos_por_estilo = fotos_por_estilo

    st.markdown("**Etiquetas por pedido (1 + 2 + 3, ZPL y PDF):**")
    if "zpl_cache" not in st.session_state:
        st.session_state.zpl_cache = {}
    if "pdf_cache" not in st.session_state:
        st.session_state.pdf_cache = {}
    if "zpl_qr_cache" not in st.session_state:
        st.session_state.zpl_qr_cache = {}
    if "pdf_qr_cache" not in st.session_state:
        st.session_state.pdf_qr_cache = {}
    if "zpl_fotos_cache" not in st.session_state:
        st.session_state.zpl_fotos_cache = {}
    if "pdf_fotos_cache" not in st.session_state:
        st.session_state.pdf_fotos_cache = {}

    for oc in pedidos:
        n_cajas_oc = len([b for b in boxes if b["oc"] == oc])
        with st.expander(f"OC {oc} — {n_cajas_oc} cajas", expanded=True):
            st.caption("Completo (Etiqueta 1 + 2 + 3):")
            colz, colp = st.columns(2)
            with colz:
                if oc not in st.session_state.zpl_cache:
                    if st.button(f"Generar ZPL — {oc}", key=f"zplbtn_{oc}"):
                        with st.spinner("Generando ZPL (incluye los QR, puede tardar)..."):
                            st.session_state.zpl_cache[oc] = eng.generar_zpl_para_oc(boxes, oc, temporada, empresa)
                        st.rerun()
                else:
                    st.download_button("⬇️ Descargar ZPL", st.session_state.zpl_cache[oc],
                                        file_name=f"Etiquetas_OC_{oc}.zpl",
                                        mime="text/plain", key=f"zpldl_{oc}")
            with colp:
                if oc not in st.session_state.pdf_cache:
                    if st.button(f"Generar PDF — {oc}", key=f"pdfbtn_{oc}"):
                        with st.spinner("Generando PDF (incluye los QR, puede tardar)..."):
                            st.session_state.pdf_cache[oc] = eng.generar_pdf_para_oc(boxes, oc, temporada, empresa)
                        st.rerun()
                else:
                    st.download_button("⬇️ Descargar PDF", st.session_state.pdf_cache[oc],
                                        file_name=f"Etiquetas_OC_{oc}.pdf",
                                        mime="application/pdf", key=f"pdfdl_{oc}")

            st.caption("Solo Etiqueta 3 (QR):")
            colz2, colp2 = st.columns(2)
            with colz2:
                if oc not in st.session_state.zpl_qr_cache:
                    if st.button(f"Generar ZPL solo QR — {oc}", key=f"zplqrbtn_{oc}"):
                        with st.spinner("Generando QR (puede tardar)..."):
                            st.session_state.zpl_qr_cache[oc] = eng.generar_zpl_solo_qr_para_oc(boxes, oc, empresa)
                        st.rerun()
                else:
                    st.download_button("⬇️ Descargar ZPL (solo QR)", st.session_state.zpl_qr_cache[oc],
                                        file_name=f"Etiquetas_QR_OC_{oc}.zpl",
                                        mime="text/plain", key=f"zplqrdl_{oc}")
            with colp2:
                if oc not in st.session_state.pdf_qr_cache:
                    if st.button(f"Generar PDF solo QR — {oc}", key=f"pdfqrbtn_{oc}"):
                        with st.spinner("Generando QR (puede tardar)..."):
                            st.session_state.pdf_qr_cache[oc] = eng.generar_pdf_solo_qr_para_oc(boxes, oc, empresa)
                        st.rerun()
                else:
                    st.download_button("⬇️ Descargar PDF (solo QR)", st.session_state.pdf_qr_cache[oc],
                                        file_name=f"Etiquetas_QR_OC_{oc}.pdf",
                                        mime="application/pdf", key=f"pdfqrdl_{oc}")

            st.caption("Etiqueta QR + fotos por modelo:")
            if not st.session_state.get("fotos_por_estilo"):
                st.caption("Sube al menos una fotografía y asígnala a un modelo arriba (paso 5) para habilitar esta opción.")
            else:
                colz3, colp3 = st.columns(2)
                with colz3:
                    if oc not in st.session_state.zpl_fotos_cache:
                        if st.button(f"Generar ZPL con fotos — {oc}", key=f"zplftbtn_{oc}"):
                            with st.spinner("Generando QR + fotos (puede tardar)..."):
                                from io import BytesIO as _BIO
                                from PIL import Image as _Image
                                fotos_img = {e: _Image.open(_BIO(fb))
                                             for e, fb in st.session_state.fotos_por_estilo.items()}
                                st.session_state.zpl_fotos_cache[oc] = eng.generar_zpl_qr_fotos_para_oc(
                                    boxes, oc, empresa, fotos_img)
                            st.rerun()
                    else:
                        st.download_button("⬇️ Descargar ZPL (QR + fotos)", st.session_state.zpl_fotos_cache[oc],
                                            file_name=f"Etiquetas_QRFotos_OC_{oc}.zpl",
                                            mime="text/plain", key=f"zplftdl_{oc}")
                with colp3:
                    if oc not in st.session_state.pdf_fotos_cache:
                        if st.button(f"Generar PDF con fotos — {oc}", key=f"pdfftbtn_{oc}"):
                            with st.spinner("Generando QR + fotos (puede tardar)..."):
                                st.session_state.pdf_fotos_cache[oc] = eng.generar_pdf_qr_fotos_para_oc(
                                    boxes, oc, empresa, st.session_state.fotos_por_estilo)
                            st.rerun()
                    else:
                        st.download_button("⬇️ Descargar PDF (QR + fotos)", st.session_state.pdf_fotos_cache[oc],
                                            file_name=f"Etiquetas_QRFotos_OC_{oc}.pdf",
                                            mime="application/pdf", key=f"pdfftdl_{oc}")
